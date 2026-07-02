"""Runway — Layer 1 engine (deterministic, NO LLM).

Reads raw DOL OFLC LCA disclosure data and produces a grounded list of
employers that have *certified* an entry-wage (PW_WAGE_LEVEL = I) filing for a
given role's SOC codes. This is the engine the future self-serve product calls;
the notebook / report scripts are thin callers (engine/altitude split).

Forward-compat choices (the two the design doc allows):
  1. Role is NOT hardcoded — it's a SOC list, seeded design-only via ROLE_SOC.
  2. Engine is separable from the report — this module returns DataFrames; it
     writes no report and imports nothing about HTML/PDF.

Adding "consultant" later = one entry in ROLE_SOC. Nothing else changes.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

# --------------------------------------------------------------------------- #
# Role -> SOC mapping.  Seeded design-only (premise: differentiator is the
# entry-level design sponsor list).  2018 SOC codes:
#   15-1255  Web and Digital Interface Designers   (the UI/UX code — primary)
#   27-1024  Graphic Designers
#   27-1021  Commercial and Industrial Designers
# Pre-2018 fallbacks (only if a quarter predates the 2018 SOC system): 15-1199,
# 27-1014.  v1 uses 2021+ quarters, so the fallbacks are documented, not wired.
# --------------------------------------------------------------------------- #
ROLE_SOC: dict[str, list[str]] = {
    "design": ["15-1255", "27-1024", "27-1021"],
}

# Human-readable SOC titles, for report context (DOL's own titles vary in case).
SOC_TITLES: dict[str, str] = {
    "15-1255": "Web and Digital Interface Designers",
    "27-1024": "Graphic Designers",
    "27-1021": "Commercial and Industrial Designers",
}

# Columns we keep from the ~98-col raw file.  Selected by NAME, resolved against
# each file's own header row — column *order is not stable across quarters*
# (e.g. FY2025 Q1 has WORKSITE_COUNTY where Q4 has WORKSITE_CITY), so hardcoded
# indices silently grab the wrong column.  Kept narrow on purpose: the file is
# 100+ MB and we never read all 98 columns into memory.
REQUIRED_COLUMNS: list[str] = [
    "CASE_STATUS",
    "VISA_CLASS",
    "JOB_TITLE",
    "SOC_CODE",
    "SOC_TITLE",
    "EMPLOYER_NAME",
    "WORKSITE_CITY",
    "WORKSITE_STATE",
    "WAGE_RATE_OF_PAY_FROM",
    "WAGE_RATE_OF_PAY_TO",
    "WAGE_UNIT_OF_PAY",
    "PW_WAGE_LEVEL",
]
_NUMERIC_COLUMNS = {"WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_TO"}

# PW_WAGE_LEVEL spelling varies by vintage: "I" / "Level I" / 1.  Run
# value_counts() before trusting this on a new quarter (FY2025 uses bare "I").
LEVEL_I_VALUES: set[str] = {"I", "LEVEL I", "1"}

# Corporate suffixes stripped when normalizing employer names so a multi-entity
# employer collapses to one row.  Conservative on purpose — over-stripping
# merges genuinely different companies.
_CORP_SUFFIXES: set[str] = {
    "LLC", "L L C", "INC", "CORP", "CORPORATION", "LTD", "LIMITED",
    "CO", "COMPANY", "LP", "LLP", "PC", "PLLC", "INCORPORATED",
}

# Wage-unit -> annual multiplier, for comparing WAGE_RATE_OF_PAY across filings.
_ANNUALIZE: dict[str, float] = {
    "YEAR": 1.0, "HOUR": 2080.0, "MONTH": 12.0, "WEEK": 52.0,
    "BI-WEEKLY": 26.0, "BIWEEKLY": 26.0,
}


# --------------------------------------------------------------------------- #
# Normalization
# --------------------------------------------------------------------------- #
def normalize_soc(code) -> str:
    """'15-1255.00' / '15-1255' -> '15-1255'.  Empty for missing."""
    if code is None:
        return ""
    s = str(code).strip()
    return s.split(".")[0][:7]


def normalize_employer(name) -> str | None:
    """Uppercase, strip punctuation + trailing corporate suffixes + leading THE.

    'The Deloitte Consulting, LLP.' -> 'DELOITTE CONSULTING'.  Returns None for
    missing names so they group together and can be excluded.
    """
    if name is None:
        return None
    s = str(name).upper()
    s = re.sub(r"[.,]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    tokens = s.split()
    while tokens and tokens[-1] in _CORP_SUFFIXES:
        tokens.pop()
    if tokens and tokens[0] == "THE":
        tokens = tokens[1:]
    return " ".join(tokens) if tokens else s


def _annualized_wage(row) -> float | None:
    val = row.get("WAGE_RATE_OF_PAY_FROM")
    if val is None or pd.isna(val):
        return None
    unit = str(row.get("WAGE_UNIT_OF_PAY") or "").upper().strip()
    mult = _ANNUALIZE.get(unit)
    if mult is None:
        return None
    return float(val) * mult


# --------------------------------------------------------------------------- #
# Raw file -> parquet (slow step, done once per quarter)
# --------------------------------------------------------------------------- #
def quarter_parquet_path(quarter: str, processed_dir: str | Path = "data/processed") -> Path:
    """'FY2025Q4' -> data/processed/lca_fy2025q4.parquet"""
    return Path(processed_dir) / f"lca_{quarter.lower()}.parquet"


_QUARTER_RE = re.compile(r"lca_(fy\d{4}q\d)\.parquet$", re.IGNORECASE)


def detect_quarters(processed_dir: str | Path = "data/processed") -> list[str]:
    """Every quarter that has actually been converted, e.g. ['FY2025Q4'].

    The tool runs on whatever data is present — one quarter is a valid run.
    Returned sorted (chronological, since the labels sort naturally), uppercased
    to the FY####Q# form the rest of the engine uses. Empty list = no data yet.
    """
    d = Path(processed_dir)
    if not d.exists():
        return []
    out = []
    for p in d.glob("lca_*.parquet"):
        m = _QUARTER_RE.search(p.name)
        if m:
            out.append(m.group(1).upper())
    return sorted(out)


def xlsx_to_parquet(xlsx_path: str | Path, parquet_path: str | Path) -> int:
    """Stream the 100+ MB raw xlsx -> narrow parquet, only REQUIRED_COLUMNS.

    Resolves columns by NAME against the file's own header (order is not stable
    across quarters).  Uses openpyxl read-only streaming (never materializes all
    ~98 cols).  Coerces wage columns to numeric and everything else to string so
    parquet is happy despite the source's mixed cell types.  Returns rows written.
    """
    import openpyxl

    xlsx_path, parquet_path = Path(xlsx_path), Path(parquet_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        name_to_idx = {h: i for i, h in enumerate(header)}
        missing = [c for c in REQUIRED_COLUMNS if c not in name_to_idx]
        if missing:
            raise ValueError(
                f"{xlsx_path.name} missing required columns {missing}. "
                f"Wrong file (PERM not LCA?) or renamed columns."
            )
        names = REQUIRED_COLUMNS
        idx = [name_to_idx[c] for c in names]
        rows = [tuple(r[i] for i in idx) for r in ws.iter_rows(min_row=2, values_only=True)]
    finally:
        wb.close()

    df = pd.DataFrame(rows, columns=names)
    for c in names:
        if c in _NUMERIC_COLUMNS:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        else:
            df[c] = df[c].map(lambda v: None if v is None else str(v))

    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(parquet_path, index=False)
    return len(df)


# --------------------------------------------------------------------------- #
# The engine
# --------------------------------------------------------------------------- #
def load_certified_rows(
    soc_codes: list[str],
    wage_level: str,
    quarters: str | list[str],
    processed_dir: str | Path = "data/processed",
) -> pd.DataFrame:
    """Row-level certified filings matching role SOCs + wage level, across quarters.

    Returns one row per matching certified filing, with a QUARTER column and
    derived EMP_NORM / SOC7 / WAGE_ANNUAL columns.  This is the grounding the
    gap-read consumes; build_sponsor_table aggregates it to one row per employer.
    """
    if isinstance(quarters, str):
        quarters = [quarters]
    soc_set = {normalize_soc(c) for c in soc_codes}
    level_set = {wage_level.upper().strip()} if isinstance(wage_level, str) else set()
    if wage_level == "I":  # accept all known spellings of Level I
        level_set = LEVEL_I_VALUES

    frames = []
    for q in quarters:
        path = quarter_parquet_path(q, processed_dir)
        if not path.exists():
            raise FileNotFoundError(f"Missing parquet for {q}: {path} (run xlsx_to_parquet)")
        df = pd.read_parquet(path)

        # Column-present assert (catches a wrong/PERM file before any filtering).
        missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"{path.name} missing required columns: {missing}")

        df["SOC7"] = df["SOC_CODE"].map(normalize_soc)
        cert = df[df["CASE_STATUS"].fillna("").str.upper() == "CERTIFIED"].copy()
        lvl = cert["PW_WAGE_LEVEL"].fillna("").str.upper().str.strip()
        sel = cert[cert["SOC7"].isin(soc_set) & lvl.isin(level_set)].copy()
        sel["QUARTER"] = q
        frames.append(sel)

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if out.empty:
        return out
    out["EMP_NORM"] = out["EMPLOYER_NAME"].map(normalize_employer)
    out = out[out["EMP_NORM"].notna()].copy()
    out["WAGE_ANNUAL"] = out.apply(_annualized_wage, axis=1)
    return out


def build_sponsor_table(
    soc_codes: list[str],
    wage_level: str,
    quarter: str | list[str],
    processed_dir: str | Path = "data/processed",
) -> pd.DataFrame:
    """Grounded entry-wage sponsor table: one row per normalized employer.

    Filing count = certified rows grouped by normalized employer name.  Sorted
    by quarters_present (repeat-sponsor = strongest "actually sponsors" signal),
    then filing_count.  `quarter` may be one label or a list (full-year roll-up).
    """
    rows = load_certified_rows(soc_codes, wage_level, quarter, processed_dir)
    if rows.empty:
        return pd.DataFrame(
            columns=[
                "employer", "employer_name_raw", "filing_count", "quarters_present",
                "quarters", "soc_codes", "soc_titles", "worksite_states",
                "worksite_cities", "wage_annual_min", "wage_annual_median",
                "wage_annual_max",
            ]
        )

    def _uniq_join(series, sep=" | ", top=None):
        vals = [str(v) for v in series.dropna().unique() if str(v).strip()]
        vals = sorted(vals)
        if top:
            vals = vals[:top]
        return sep.join(vals)

    def _soc_codes(series):
        return ", ".join(sorted(series.dropna().unique()))

    def _soc_titles(series):
        codes = sorted(series.dropna().unique())
        return ", ".join(SOC_TITLES.get(c, c) for c in codes)

    grouped = rows.groupby("EMP_NORM")
    table = grouped.agg(
        filing_count=("EMPLOYER_NAME", "size"),
        quarters_present=("QUARTER", "nunique"),
        quarters=("QUARTER", lambda s: ", ".join(sorted(s.unique()))),
        soc_codes=("SOC7", _soc_codes),
        soc_titles=("SOC7", _soc_titles),
        worksite_states=("WORKSITE_STATE", lambda s: _uniq_join(s)),
        worksite_cities=("WORKSITE_CITY", lambda s: _uniq_join(s, top=8)),
        wage_annual_min=("WAGE_ANNUAL", "min"),
        wage_annual_median=("WAGE_ANNUAL", "median"),
        wage_annual_max=("WAGE_ANNUAL", "max"),
        employer_name_raw=("EMPLOYER_NAME", lambda s: s.mode().iat[0] if not s.mode().empty else s.iat[0]),
    ).reset_index().rename(columns={"EMP_NORM": "employer"})

    table = table.sort_values(
        ["quarters_present", "filing_count", "employer"],
        ascending=[False, False, True],
    ).reset_index(drop=True)

    # round wages for readability
    for c in ("wage_annual_min", "wage_annual_median", "wage_annual_max"):
        table[c] = table[c].round(0)

    cols = [
        "employer", "employer_name_raw", "filing_count", "quarters_present",
        "quarters", "soc_codes", "soc_titles", "worksite_states",
        "worksite_cities", "wage_annual_min", "wage_annual_median", "wage_annual_max",
    ]
    return table[cols]
